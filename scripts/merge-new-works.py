#!/usr/bin/env python3
"""Merge new works from the intake tool (Werke_neu.csv + photos) into the site + inventory.

This is the bridge from Ute's 'Werk-Erfassung.html' export to the live website.

Usage:
  python3 scripts/merge-new-works.py <Werke_neu.csv> <ordner-mit-fotos>

Does, for each row:
  - copies the photo ('<Titel> - <MG-Nr>.<ext>') into public/images/artworks/<slug>__0.<ext>
    and into '05 Strategie/Inventar/Bilder/'
  - appends a new artwork to data/artworks.json  (what the website renders)
  - appends a row to the master inventory (xlsx + csv)
Backs up artworks.json and the master before writing. Skips rows whose Nr/slug
already exist. After running, deploy with:  npm run deploy
"""

import csv
import json
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
ART = ROOT / "data" / "artworks.json"
CATS = ROOT / "data" / "categories.json"
PUB = ROOT / "public" / "images" / "artworks"
INV_DIR = ROOT.parent / "05 Strategie" / "Inventar"
MASTER_X = INV_DIR / "Mauckert_Inventar_master.xlsx"
MASTER_C = INV_DIR / "Mauckert_Inventar_master.csv"
BILDER = INV_DIR / "Bilder"
BACKUP = INV_DIR / "Backup"


def slugify(s):
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9äöüß\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    return re.sub(r"-+", "-", s).strip("-")


def safe(s):
    s = unicodedata.normalize("NFKC", s or "")
    return re.sub(r"[\\/:*?\"<>|]+", "-", s).strip().rstrip(".")


def de_money(v):
    s = f"{v:,.2f}".replace(",", "§").replace(".", ",").replace("§", ".")
    return s + "€"


def num(s):
    try:
        return int(float(str(s).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    csv_path, img_dir = Path(sys.argv[1]), Path(sys.argv[2])
    rows = list(csv.DictReader(csv_path.open(encoding="utf-8-sig")))
    cat_map = {c["name"]: c["slug"] for c in json.loads(CATS.read_text())}
    artworks = json.loads(ART.read_text(encoding="utf-8"))
    have_slugs = {a["slug"] for a in artworks}

    # backups
    BACKUP.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(ART, ART.with_name(f"artworks.backup-{ts}.json"))
    shutil.copy2(MASTER_X, BACKUP / f"Mauckert_Inventar_master_{ts}.xlsx")
    shutil.copy2(MASTER_C, BACKUP / f"Mauckert_Inventar_master_{ts}.csv")

    wb = openpyxl.load_workbook(MASTER_X)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    have_nr = {ws.cell(r, headers.index("Nr") + 1).value for r in range(2, ws.max_row + 1)}

    added, skipped = 0, 0
    for r in rows:
        nr = (r.get("Nr") or "").strip()
        title = (r.get("Titel") or "").strip()
        if not title:
            continue
        slug = slugify(title)
        if slug in have_slugs or nr in have_nr:
            print(f"  übersprungen (existiert schon): {nr} {title}")
            skipped += 1
            continue

        techs = [t.strip() for t in (r.get("Technik") or "").split(",") if t.strip()]
        cat_slugs = [cat_map.get(t, slugify(t)) for t in techs]
        price = num(r.get("Preis (€)"))
        status = (r.get("Status") or "Verfügbar").strip()
        beschr = (r.get("Beschreibung") or "").strip()

        # image
        local_path = ""
        matches = sorted(img_dir.glob(f"* - {nr}.*")) or sorted(img_dir.glob(f"*{slug}*.*"))
        if matches:
            src = matches[0]
            ext = src.suffix.lower()
            PUB.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, PUB / f"{slug}__0{ext}")
            BILDER.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, BILDER / f"{safe(title)} - {nr}{ext}")
            local_path = f"/images/artworks/{slug}__0{ext}"
        else:
            print(f"  ! kein Foto gefunden für {nr} {title} (Werk trotzdem angelegt)")

        # website artwork
        artworks.append({
            "slug": slug, "name": title,
            "description": f"<p>{beschr}</p>" if beschr else "",
            "price": price, "currency": "EUR",
            "formatted_price": de_money(price) if price is not None else None,
            "category_ids": [], "categories": techs, "category_slugs": cat_slugs,
            "in_stock": status != "Verkauft",
            "media": [{"id": f"{slug}__0", "title": title, "width": None, "height": None,
                       "original_url": "", "display_url": local_path, "local_path": local_path}] if local_path else [],
            "source_url": "",
        })
        have_slugs.add(slug)

        # master inventory row
        gaps = [g for g, ok in (("Jahr", r.get("Jahr")), ("Maße", r.get("Maße (cm)")),
                                ("Materialkosten", r.get("Materialkosten (€)"))) if not (ok or "").strip()]
        rowmap = {
            "Nr": nr, "WV-Nr": "", "Titel": title, "Jahr": r.get("Jahr", ""),
            "Technik": ", ".join(techs), "Kategorie": ", ".join(techs),
            "Maße (cm)": r.get("Maße (cm)", ""), "Materialkosten (€)": r.get("Materialkosten (€)", ""),
            "Arbeitsstunden": r.get("Arbeitsstunden", ""), "Preis (€)": price if price is not None else "",
            "Status": status, "Käufer / Notiz": r.get("Käufer / Notiz", ""),
            "Gerahmt": r.get("Gerahmt", ""), "Masterpiece": "Nein", "Auf Website": "Ja",
            "Slug": slug, "Bild": local_path, "Datenlücken": ", ".join(gaps),
        }
        ws.append([rowmap.get(h, "") for h in headers])
        added += 1
        print(f"  ✓ {nr} {title}  ({status}, {price or '—'}€)")

    # write site + master
    ART.write_text(json.dumps(artworks, ensure_ascii=False, indent=2), encoding="utf-8")
    wb.save(MASTER_X)
    wb2 = openpyxl.load_workbook(MASTER_X, data_only=True)
    ws2 = wb2.active
    with MASTER_C.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws2.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])

    print(f"\n{added} Werke hinzugefügt, {skipped} übersprungen.")
    print("Backups in Inventar/Backup/ + data/artworks.backup-*.json")
    print("→ Live schalten:  cd website && npm run deploy")


if __name__ == "__main__":
    main()
