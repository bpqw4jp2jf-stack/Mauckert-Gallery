#!/usr/bin/env python3
"""Merge Wix-enriched artworks with the Werkverzeichnis and emit an Excel inventory.

Output columns:
  - Nr (Werkverzeichnis number if matched)
  - Titel
  - Jahr (creation year, best-effort)
  - Technik / Medium
  - Maße
  - Preis (EUR)
  - Verkauft (Ja/Nein)
  - Masterpiece (Ja/Nein)        – "Hommage an …" works
  - Gerahmt (Ja/Nein/Unbekannt)  – only when explicitly mentioned
  - Kategorien
  - Beschreibung
  - Website-Slug
  - Bild (local path)
  - Quelle (URL)
"""

import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
ENRICHED = ROOT / "data" / "artworks_enriched.json"
WV = ROOT / "data" / "werkverzeichnis.json"
OUT = ROOT / "data" / "Mauckert_Gallery_Inventar.xlsx"

NAVY = "FF101C28"
CREAM = "FFF7F0D8"
STONE = "FFC7BFA8"


def norm(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def kw_overlap(a: str, b: str) -> int:
    aw = {w for w in norm(a).split() if len(w) >= 4}
    bw = {w for w in norm(b).split() if len(w) >= 4}
    return len(aw & bw)


def find_wv_match(title: str, wv: list[dict]) -> dict | None:
    n = norm(title)
    # Exact
    for r in wv:
        if norm(r.get("titel")) == n:
            return r
    # Best keyword overlap (>= 2 words)
    best = (0, None)
    for r in wv:
        score = kw_overlap(title, r.get("titel") or "")
        if score > best[0]:
            best = (score, r)
    if best[0] >= 2:
        return best[1]
    return None


def yn(b) -> str:
    if b is True:
        return "Ja"
    if b is False:
        return "Nein"
    return "Unbekannt"


def main() -> None:
    artworks = json.loads(ENRICHED.read_text())
    wv = json.loads(WV.read_text())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventar"

    headers = [
        "Nr (WV)", "Titel", "Jahr", "Technik / Medium", "Maße", "Preis (EUR)",
        "Verkauft", "Masterpiece", "Gerahmt", "Ribbon (Wix-Hinweis)",
        "Kategorien", "Beschreibung", "Slug", "Bild", "Quelle",
    ]
    ws.append(headers)

    head_font = Font(color="FFFFFFFF", bold=True, name="Calibri")
    head_fill = PatternFill("solid", fgColor=NAVY)
    head_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = head_font
        c.fill = head_fill
        c.alignment = head_align
    ws.row_dimensions[1].height = 28

    row_num = 1
    BUYER_RIBBONS = {"amf capital", "lume", "kuku vaia"}
    for a in artworks:
        e = a.get("enrichment") or {}
        ribbon = (e.get("ribbon") or "").strip()
        rb_lower = ribbon.lower()
        match = find_wv_match(a["name"], wv)
        # Sold: explicit Wix "Verkauft", any buyer ribbon, out-of-stock, sold-hint, or WV "verkauft" status
        is_sold = (
            rb_lower == "verkauft"
            or any(b in rb_lower for b in BUYER_RIBBONS)
            or e.get("isInStock") is False
            or bool(e.get("sold_hint"))
            or (match and "verkauf" in (match.get("status") or "").lower())
        )
        # Masterpiece: explicit Wix ribbon, OR title contains MP/Hommage/Plate
        is_mp = (
            rb_lower == "masterpiece"
            or bool(e.get("mp_hint"))
            or "plate" in norm(a["name"])
        )
        # Gerahmt: only set when explicit hint
        gerahmt = True if e.get("framed_hint") else None
        # Year: prefer WV, fall back to enrichment
        year = None
        if match and match.get("jahr"):
            try: year = int(match["jahr"])
            except: year = match["jahr"]
        if not year and e.get("year_guess"):
            year = e["year_guess"]
        # Maße
        size = (match.get("masse") if match else None) or e.get("size_guess") or ""
        # Technik
        technik = (match.get("technik") if match else None) or ", ".join(a.get("categories") or [])
        # Beschreibung
        beschreibung = (e.get("description_text") or "").strip()

        row = [
            (match.get("nr") if match else "") or "",
            a["name"],
            year or "",
            technik,
            size,
            a.get("price") or "",
            yn(is_sold),
            yn(is_mp),
            yn(gerahmt) if gerahmt is not None else "Unbekannt",
            ribbon,
            ", ".join(a.get("categories") or []),
            beschreibung,
            a.get("slug") or "",
            (a.get("media") or [{}])[0].get("local_path", ""),
            a.get("source_url") or "",
        ]
        ws.append(row)
        row_num += 1

    # Column widths
    widths = [10, 38, 8, 28, 16, 12, 11, 13, 12, 20, 22, 50, 32, 32, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap text on long columns
    for col_idx in (2, 4, 11, 12, 14, 15):
        for r in range(2, row_num + 1):
            ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze header
    ws.freeze_panes = "A2"

    # Conditional shading for Verkauft / Masterpiece columns
    sold_fill = PatternFill("solid", fgColor="FFEFD9CF")  # soft clay
    mp_fill = PatternFill("solid", fgColor=CREAM)
    for r in range(2, row_num + 1):
        if ws.cell(row=r, column=7).value == "Ja":
            ws.cell(row=r, column=7).fill = sold_fill
        if ws.cell(row=r, column=8).value == "Ja":
            ws.cell(row=r, column=8).fill = mp_fill

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}  ({row_num - 1} Werke)")


if __name__ == "__main__":
    main()
