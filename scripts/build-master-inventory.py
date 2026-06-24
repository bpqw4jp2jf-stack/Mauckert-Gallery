#!/usr/bin/env python3
"""Build the single source-of-truth inventory for Mauckert Gallery.

Spine = the 132 works on the website (data/artworks_enriched.json), because that
is the most complete list. Each work is enriched with the official
Werkverzeichnis (data/werkverzeichnis.json) where a title match exists.

Outputs (into ../05 Strategie/Inventar/):
  - Mauckert_Inventar_master.xlsx   formatted, the file Ute maintains by hand
  - Mauckert_Inventar_master.csv    same data, diff-friendly / re-importable
  - Datenluecken.md                 gap report: what is still missing per work

After the first run this becomes a HAND-MAINTAINED file. Re-running overwrites
it, so only re-run when you want to rebuild from the website + Werkverzeichnis.
"""

import csv
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
ENRICHED = ROOT / "data" / "artworks_enriched.json"
WV = ROOT / "data" / "werkverzeichnis.json"
OUT_DIR = ROOT.parent / "05 Strategie" / "Inventar"
XLSX = OUT_DIR / "Mauckert_Inventar_master.xlsx"
CSV = OUT_DIR / "Mauckert_Inventar_master.csv"
GAPS = OUT_DIR / "Datenluecken.md"

# Brand palette (matches website src/styles/global.css)
NAVY = "FF101C28"
CREAM = "FFF7F0D8"
CLAY = "FFEFD9CF"
SAGE = "FFE3E8DD"

STATUS_OPTIONS = ["Verfügbar", "Verkauft", "Reserviert", "Ausgestellt"]

HEADERS = [
    "Nr", "WV-Nr", "Titel", "Jahr", "Technik", "Kategorie", "Maße (cm)",
    "Materialkosten (€)", "Arbeitsstunden", "Preis (€)", "Status", "Käufer / Notiz",
    "Gerahmt", "Masterpiece", "Auf Website", "Slug", "Bild", "Datenlücken",
]


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def kw_overlap(a, b):
    aw = {w for w in norm(a).split() if len(w) >= 4}
    bw = {w for w in norm(b).split() if len(w) >= 4}
    return len(aw & bw)


def find_wv_match(title, wv):
    n = norm(title)
    for r in wv:
        if norm(r.get("titel")) == n:
            return r
    best = (0, None)
    for r in wv:
        score = kw_overlap(title, r.get("titel") or "")
        if score > best[0]:
            best = (score, r)
    return best[1] if best[0] >= 2 else None


def yn(b):
    return "Ja" if b is True else ("Nein" if b is False else "")


def clean_size(raw):
    """Normalise '0.30 x 0,30' / '30x40 cm' to 'AA x BB' in cm, best-effort."""
    if not raw:
        return ""
    s = raw.replace(",", ".")
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if len(nums) >= 2:
        def to_cm(v):
            f = float(v)
            if f < 5:  # metres -> cm
                f *= 100
            return str(int(round(f)))
        return f"{to_cm(nums[0])} x {to_cm(nums[1])}"
    return raw.strip()


BUYER_RIBBONS = {"amf capital", "lume", "kuku vaia"}


def build_rows():
    artworks = json.loads(ENRICHED.read_text())
    wv = json.loads(WV.read_text())

    rows = []
    for idx, a in enumerate(artworks, 1):
        e = a.get("enrichment") or {}
        match = find_wv_match(a["name"], wv)
        ribbon = (e.get("ribbon") or "").strip()
        rb = ribbon.lower()

        # --- Status ---
        wv_status = (match.get("status") or "").lower() if match else ""
        buyer = ""
        is_sold = (
            rb == "verkauft"
            or any(b in rb for b in BUYER_RIBBONS)
            or e.get("isInStock") is False
            or bool(e.get("sold_hint"))
            or "verkauf" in wv_status
        )
        status = "Verkauft" if is_sold else "Verfügbar"
        if match and "verkauft" in wv_status:
            # capture buyer name after the word "verkauft"
            buyer = re.sub(r".*verkauft", "", match.get("status"), flags=re.I).strip()
        elif any(b in rb for b in BUYER_RIBBONS):
            buyer = ribbon

        # --- Nr: unique primary key per work; WV-Nr = official ref (may be blank/repeat) ---
        nr = f"MG-{idx:03d}"
        wv_nr = (match.get("nr") if match else None) or ""

        # --- Year ---
        year = ""
        if match and match.get("jahr"):
            year = match["jahr"]
        elif e.get("year_guess"):
            year = e["year_guess"]

        # --- Size ---
        size = clean_size((match.get("masse") if match else None) or e.get("size_guess"))

        # --- Technik ---
        technik = (match.get("technik") if match else None) or ", ".join(a.get("categories") or [])
        kategorie = ", ".join(a.get("categories") or [])

        # --- Price ---
        preis = a.get("price") or (match.get("preis") if match else None) or ""

        is_mp = rb == "masterpiece" or bool(e.get("mp_hint")) or "plate" in norm(a["name"])
        gerahmt = True if e.get("framed_hint") else None

        # --- Data gaps ---
        gaps = []
        if not year:
            gaps.append("Jahr")
        if not size:
            gaps.append("Maße")
        gaps.append("Materialkosten")  # cost is nowhere yet, always to fill
        if not preis:
            gaps.append("Preis")

        rows.append({
            "Nr": nr,
            "WV-Nr": wv_nr,
            "Titel": a["name"],
            "Jahr": year,
            "Technik": technik,
            "Kategorie": kategorie,
            "Maße (cm)": size,
            "Materialkosten (€)": "",
            "Arbeitsstunden": "",
            "Preis (€)": preis,
            "Status": status,
            "Käufer / Notiz": buyer,
            "Gerahmt": yn(gerahmt),
            "Masterpiece": yn(is_mp),
            "Auf Website": "Ja",
            "Slug": a.get("slug") or "",
            "Bild": (a.get("media") or [{}])[0].get("local_path", ""),
            "Datenlücken": ", ".join(gaps),
        })
    return rows


def write_csv(rows):
    with CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventar"
    ws.append(HEADERS)

    head_font = Font(color="FFFFFFFF", bold=True)
    head_fill = PatternFill("solid", fgColor=NAVY)
    head_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFD8D2C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font, c.fill, c.alignment = head_font, head_fill, head_align
    ws.row_dimensions[1].height = 30

    for r in rows:
        ws.append([r[h] for h in HEADERS])

    last = len(rows) + 1

    # Status dropdown
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS_OPTIONS), allow_blank=False)
    ws.add_data_validation(dv)
    status_col = HEADERS.index("Status") + 1
    dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{last}")

    widths = [9, 10, 36, 7, 26, 18, 12, 16, 13, 11, 13, 22, 9, 12, 11, 30, 30, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap_cols = {HEADERS.index(h) + 1 for h in ("Titel", "Technik", "Kategorie", "Slug", "Bild", "Datenlücken")}
    sold_fill = PatternFill("solid", fgColor=CLAY)
    mp_fill = PatternFill("solid", fgColor=CREAM)
    gap_font = Font(color="FFB23A2E", italic=True)
    for ri in range(2, last + 1):
        for ci in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            if ci in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        if ws.cell(row=ri, column=HEADERS.index("Status") + 1).value == "Verkauft":
            ws.cell(row=ri, column=HEADERS.index("Status") + 1).fill = sold_fill
        if ws.cell(row=ri, column=HEADERS.index("Masterpiece") + 1).value == "Ja":
            ws.cell(row=ri, column=HEADERS.index("Masterpiece") + 1).fill = mp_fill
        ws.cell(row=ri, column=HEADERS.index("Datenlücken") + 1).font = gap_font

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last}"
    wb.save(XLSX)


def write_gaps(rows):
    n = len(rows)
    miss_year = [r for r in rows if "Jahr" in r["Datenlücken"]]
    miss_size = [r for r in rows if "Maße" in r["Datenlücken"]]
    miss_price = [r for r in rows if "Preis" in r["Datenlücken"]]
    sold = [r for r in rows if r["Status"] == "Verkauft"]

    lines = [
        "# Datenlücken im Inventar",
        "",
        f"Stand: automatisch erzeugt aus Website + Werkverzeichnis. **{n} Werke** gesamt.",
        "",
        "## Überblick",
        "",
        f"- Ohne **Jahr**: {len(miss_year)} / {n}",
        f"- Ohne **Maße**: {len(miss_size)} / {n}",
        f"- Ohne **Materialkosten**: {n} / {n}  (existiert noch nirgends — Pflichtfeld für die Preislogik, WP4)",
        f"- Ohne **Preis**: {len(miss_price)} / {n}",
        f"- Als **verkauft** erkannt: {len(sold)} / {n}",
        "",
        "> Materialkosten fehlen bei allen Werken, weil sie bisher nirgendwo erfasst wurden.",
        "> Sie sind die Grundlage für die Preislogik (Arbeitspaket 4).",
        "",
        "## Werke ohne Maße",
        "",
    ]
    for r in miss_size:
        lines.append(f"- {r['Nr']} — {r['Titel']}")
    lines += ["", "## Werke ohne Jahr", ""]
    for r in miss_year:
        lines.append(f"- {r['Nr']} — {r['Titel']}")
    GAPS.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    write_gaps(rows)
    sold = sum(1 for r in rows if r["Status"] == "Verkauft")
    print(f"{len(rows)} Werke → {XLSX.name} / .csv  ({sold} verkauft)")
    print(f"Gap-Report → {GAPS.name}")


if __name__ == "__main__":
    main()
