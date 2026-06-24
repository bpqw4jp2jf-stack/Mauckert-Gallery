#!/usr/bin/env python3
"""Apply the Stufe-1 prices into the master inventory's "Preis (€)" column.

- Backs up the current xlsx + csv first (timestamped, into Inventar/Backup/).
- Reads Stufe-1 prices from Preisstrategie/Preisliste_neu.csv (Nr -> price).
- Updates ONLY works that have a Stufe-1 price (those with Maße); all others
  keep their current price untouched.
- Edits the xlsx in place (formatting / dropdowns preserved) and rewrites the
  csv so both stay in sync.
- The old prices remain visible in Preisliste_neu.csv (column "Aktuell (€)").
"""

import csv
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INV = ROOT.parent / "05 Strategie" / "Inventar"
XLSX = INV / "Mauckert_Inventar_master.xlsx"
CSV = INV / "Mauckert_Inventar_master.csv"
PRICELIST = ROOT.parent / "05 Strategie" / "Preisstrategie" / "Preisliste_neu.csv"
BACKUP = INV / "Backup"


def main():
    # 1. backup
    BACKUP.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(XLSX, BACKUP / f"Mauckert_Inventar_master_{ts}.xlsx")
    shutil.copy2(CSV, BACKUP / f"Mauckert_Inventar_master_{ts}.csv")

    # 2. Stufe-1 prices by Nr
    new_price = {}
    for r in csv.DictReader(PRICELIST.open(encoding="utf-8")):
        v = r.get("Stufe 1 (€)", "").strip()
        if v:
            new_price[r["Nr"]] = int(float(v))

    # 3. edit xlsx in place
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    nr_col = headers.index("Nr") + 1
    preis_col = headers.index("Preis (€)") + 1
    status_col = headers.index("Status") + 1

    changed = skipped_sold = 0
    for row in range(2, ws.max_row + 1):
        nr = ws.cell(row=row, column=nr_col).value
        if nr not in new_price:
            continue
        # never touch sold works — they keep their historical sale price
        if ws.cell(row=row, column=status_col).value == "Verkauft":
            skipped_sold += 1
            continue
        old = ws.cell(row=row, column=preis_col).value
        if old != new_price[nr]:
            ws.cell(row=row, column=preis_col).value = new_price[nr]
            changed += 1
    wb.save(XLSX)

    # 4. rewrite csv from the (now updated) xlsx so both match
    wb2 = openpyxl.load_workbook(XLSX, data_only=True)
    ws2 = wb2.active
    rows = list(ws2.iter_rows(values_only=True))
    with CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])

    print(f"Backup → Inventar/Backup/…_{ts}.xlsx/.csv")
    print(f"{changed} Preise auf Stufe 1 aktualisiert; {skipped_sold} verkaufte Werke "
          f"unverändert gelassen ({len(new_price)} hatten einen Stufe-1-Preis).")


if __name__ == "__main__":
    main()
